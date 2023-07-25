# imports
import numpy as np
import copy
import openpyxl
import time
from openpyxl.styles import Font, Alignment


# PLEASE INPUT ALL REMAINING MATCHES HERE
MATCHES = [['RYU', 'GMY'], ['GMY', 'UTE'], ['LEM', 'RYU'], ['300', 'LEM']]


# THE CURRENT STANDING, TEAM:[WINS,LOSSES]
STANDING = {'300': [7, 1], 'GMY': [4, 2], 'RYU': [4, 2], 'LEM': [4, 2], 'DRP': [3, 7], 'UTE': [0, 8], 'PP': [0, 10], 'DSP': [0, 10]}


# THIS INDCIATES THE CHANGE OF SCORE FOR EVERY OUTCOME
# PER DEFAULT A WIN GRANTS 2 WINS AND 0 DEFEATS FOR THE WINNING TEAM
# AND 0 WINS AND 2 DEFEATS FOR THE LOOSING TEAM
DELTA_SCORE = [[[2, 0], [0, 2]], [[0, 2], [2, 0]], [[1, 1], [1, 1]]]


# THIS SPECIFIES THE TEXT FOR EVERY OUTCOME
# PER DEFAULT "0" MEANS THE FIRST TEAM WINS, THEREFORE THE STRING #0 GETS PRINTED
# STRINGS WITH # AT THE START GET TRANSLATED TO THE TEAM AT THIS POSITION
OUTCOME_DICTIONARY = {"0": "#0", "1": "#1", "2": "DRAW"}


# FILTER FOR YOUR TEAM? AND IF SO, WHAT IS YOUR TEAM
FILTER = True
MYTEAM = "RYU"


# HOW MANY TEAMS WILL REACH PLAYOFFS/RELEGATION
PLAYOFF_TEAMS = 2
RELEGATION_TEAMS = 2


# COLOR PREFERENCE
TIEBREAK_PLAYOFFS_COLOR = "3399ff"
TIEBREAK_RELEGATION_COLOR = "ff5050"


# OUTPUT FILE
OUTPUT_FILE = "./out/foldy_sheet.xlsx"

START = time.time()


# GENERATE ALL POSSIBLE OUTCOME CODES
outcome_codes = []
for code in range(len(DELTA_SCORE) ** len(MATCHES)):
    outcome_codes.append(
        str.zfill(np.base_repr(code, base=len(DELTA_SCORE)), len(MATCHES))
    )


# USE CODES AND DELTA_SCORE TO GENERATE STANDINGS
all_possible_standings = []
for code in outcome_codes:
    delta_standing = copy.deepcopy(STANDING)

    for matchIndex, matchCode in enumerate(code):
        team_A = MATCHES[matchIndex][0]
        team_B = MATCHES[matchIndex][1]
        delta_A = DELTA_SCORE[int(matchCode)][0]
        delta_B = DELTA_SCORE[int(matchCode)][1]

        delta_standing[team_A][0] += delta_A[0]
        delta_standing[team_A][1] += delta_A[1]
        delta_standing[team_B][0] += delta_B[0]
        delta_standing[team_B][1] += delta_B[1]

    delta_standing_sorted = dict(
        sorted(delta_standing.items(), key=lambda x: x[1][0], reverse=True)
    )
    all_possible_standings.append(delta_standing_sorted)


# GENERATE TIEBREAKER DATA
all_possible_standings_data = []
for standing in all_possible_standings:
    score_list = list(standing.values())
    playoffs_tiebreak_start = 0
    playoffs_tiebreak_end = PLAYOFF_TEAMS
    playoffs_tiebreak_bool = False

    relegation_tiebreak_start = 0
    relegation_tiebreak_end = len(score_list) - RELEGATION_TEAMS
    relegation_tiebreak_bool = False

    # CHECK FOR PLAYOFF TIEBREAKER
    if score_list[PLAYOFF_TEAMS] == score_list[PLAYOFF_TEAMS - 1]:
        playoffs_tiebreak_bool = True
        for score in score_list:
            if score == score_list[PLAYOFF_TEAMS]:
                break
            playoffs_tiebreak_start += 1

        for n in range(PLAYOFF_TEAMS, len(score_list)):
            if score_list[n] != score_list[PLAYOFF_TEAMS]:
                break
            playoffs_tiebreak_end += 1

        playoffs_tiebreak_end -= 1

    # CHECK FOR RELEGATION TIEBREAKER

    if (
        score_list[len(score_list) - RELEGATION_TEAMS - 1]
        == score_list[len(score_list) - RELEGATION_TEAMS]
    ):
        relegation_tiebreak_bool = True

        for score in score_list:
            if score == score_list[len(score_list) - RELEGATION_TEAMS]:
                break
            relegation_tiebreak_start += 1

        for n in range(len(score_list) - RELEGATION_TEAMS, len(score_list)):
            if score_list[n] != score_list[len(score_list) - RELEGATION_TEAMS]:
                break
            relegation_tiebreak_end += 1

        relegation_tiebreak_end -= 1

    playoff_data = [
        playoffs_tiebreak_bool,
        playoffs_tiebreak_start,
        playoffs_tiebreak_end,
    ]
    relegation_data = [
        relegation_tiebreak_bool,
        relegation_tiebreak_start,
        relegation_tiebreak_end,
    ]
    standing_data = [standing, playoff_data, relegation_data]
    all_possible_standings_data.append(standing_data)


# FILTER TIEBREAKER FOR TEAM
if FILTER:
    for standing_data_index, standing_data in enumerate(all_possible_standings_data):
        # scan for playoff importancy:
        playoff_data = standing_data[1]

        if playoff_data[0]:
            all_possible_standings_data[standing_data_index][1][0] = False
            for n in range(playoff_data[1], playoff_data[2] + 1):
                if list(standing_data[0].keys())[n] == MYTEAM:
                    all_possible_standings_data[standing_data_index][1][0] = True

        relegation_data = standing_data[2]

        if relegation_data[0]:
            all_possible_standings_data[standing_data_index][2][0] = False
            for n in range(relegation_data[1], relegation_data[2] + 1):
                if list(standing_data[0].keys())[n] == MYTEAM:
                    all_possible_standings_data[standing_data_index][2][0] = True


workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Foldy'


# HEADER: TEAM VS TEAM
for column, match in enumerate(MATCHES, start=1):
    cell = sheet.cell(row=1, column=column)
    cell.value = f"{match[0]} VS. {match[1]}"
    cell.font = Font(bold=True, underline='single')
    cell.alignment = Alignment(horizontal='center')


# HEADER: PLACES
for number in range(len(MATCHES) + 1, len(MATCHES) + 1 + len(STANDING)):
    cell = sheet.cell(row=1, column=number+1)
    cell.value = number - len(MATCHES)
    if (number - len(MATCHES)) <= PLAYOFF_TEAMS:
        cell.font = Font(underline='single', bold=True)
    cell.alignment = Alignment(horizontal='center')


# HEADER: TEAM LOCK IN?
if FILTER:
    cell = sheet.cell(row=1, column=len(MATCHES) + len(STANDING) + 3)
    cell.value = f"{MYTEAM} LOCK IN"
    cell.font = Font(bold=True, underline='single')
    cell.alignment = Alignment(horizontal='center')


# FILL OUTCOMES
for row, standing in enumerate(all_possible_standings_data):
    code = outcome_codes[row]

    for column, c in enumerate(code, start=1):
        cell = sheet.cell(row=row+3, column=column)
        cell_string = OUTCOME_DICTIONARY[c]

        if cell_string.startswith('#'):
            cell_string = cell_string[1:]
            cell_string = MATCHES[column-1][int(cell_string)]

        cell.value = cell_string
        cell.alignment = Alignment(horizontal='center')

        if cell_string == MYTEAM and FILTER:
            cell.font = Font(bold=True, underline='single')


# FILL STANDING
for row, standing in enumerate(all_possible_standings_data):
    for column, team in enumerate(standing[0].keys(), start=len(MATCHES)+2):
        cell = sheet.cell(row=row+3, column=column)
        cell.value = team
        cell.alignment = Alignment(horizontal='center')

        if team == MYTEAM and FILTER:
            cell.font = Font(bold=True, underline='single')

        index = column - len(MATCHES) - 2

        playoffs_tiebreak_color = openpyxl.styles.colors.Color(
            rgb=TIEBREAK_PLAYOFFS_COLOR
        )
        playoffs_tiebreak_fill = openpyxl.styles.fills.PatternFill(
            patternType="solid", fgColor=playoffs_tiebreak_color
        )

        relegation_tiebreak_color = openpyxl.styles.colors.Color(
            rgb=TIEBREAK_RELEGATION_COLOR
        )
        relegation_tiebreak_fill = openpyxl.styles.fills.PatternFill(
            patternType="solid", fgColor=relegation_tiebreak_color
        )

        if standing[2][0] and index >= standing[2][1] and index <= standing[2][2]:
            cell.fill = relegation_tiebreak_fill

        if standing[1][0] and index >= standing[1][1] and index <= standing[1][2]:
            cell.fill = playoffs_tiebreak_fill


# FILL TEAM LOCKIN
if FILTER:
    for row, standing in enumerate(all_possible_standings_data, start=3):
        team_list = list(standing[0].keys())
        team_index = team_list.index(MYTEAM)
        cell = sheet.cell(column=len(STANDING)+len(MATCHES)+3, row=row)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)

        playoffs_tiebreak_color = openpyxl.styles.colors.Color(
            rgb=TIEBREAK_PLAYOFFS_COLOR)
        playoffs_tiebreak_fill = openpyxl.styles.fills.PatternFill(
            patternType='solid', fgColor=playoffs_tiebreak_color)

        yes_color = openpyxl.styles.colors.Color(rgb='00cc00')
        yes_fill = openpyxl.styles.fills.PatternFill(
            patternType='solid', fgColor=yes_color)

        no_color = openpyxl.styles.colors.Color(rgb='ff0066')
        no_fill = openpyxl.styles.fills.PatternFill(
            patternType='solid', fgColor=no_color)

        if team_index < PLAYOFF_TEAMS:
            cell.value = 'YES'
            cell.fill = yes_fill
        else:
            cell.value = 'NO'
            cell.fill = no_fill

        if standing[1][0] or standing[2][0]:
            cell.value = 'TIEBREAK'
            cell.fill = playoffs_tiebreak_fill


if FILTER:
    all_scenarios = len(OUTCOME_DICTIONARY) ** len(MATCHES)
    winning_scenarios = 0
    tiebreak_scenarios = 0
    loosing_scenarios = 0

    for row in range(all_scenarios):
        column = 1 + len(MATCHES) + 1 + len(STANDING) + 1
        val = sheet.cell(row=row+3, column=column).value

        if val == 'YES':
            winning_scenarios += 1

        if val == 'NO':
            loosing_scenarios += 1

        if val == 'TIEBREAK':
            tiebreak_scenarios += 1
    workbook.create_sheet('Scenarios')
    sheet = workbook['Scenarios']
    sheet.cell(row=1, column=2).value = 'H'
    sheet.cell(row=1, column=3).value = 'h'
    sheet.cell(row=1, column=2).alignment = Alignment(horizontal='center')
    sheet.cell(row=1, column=3).alignment = Alignment(horizontal='center')
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row=1, column=3).font = Font(bold=True)
    sheet.cell(row=2, column=1).value = 'TOTAL SCENARIOS'
    sheet.cell(row=3, column=1).value = 'LOCK IN'
    sheet.cell(row=4, column=1).value = 'LOCK OUT'
    sheet.cell(row=5, column=1).value = 'TIEBREAKER'
    sheet.cell(row=2, column=1).font = Font(bold=True)
    sheet.cell(row=3, column=1).font = Font(bold=True)
    sheet.cell(row=4, column=1).font = Font(bold=True)
    sheet.cell(row=5, column=1).font = Font(bold=True)
    sheet.cell(row=2, column=2).value = all_scenarios
    sheet.cell(row=3, column=2).value = winning_scenarios
    sheet.cell(row=4, column=2).value = loosing_scenarios
    sheet.cell(row=5, column=2).value = tiebreak_scenarios
    sheet.cell(row=2, column=3).value = all_scenarios / all_scenarios
    sheet.cell(row=3, column=3).value = winning_scenarios / all_scenarios
    sheet.cell(row=4, column=3).value = loosing_scenarios / all_scenarios
    sheet.cell(row=5, column=3).value = tiebreak_scenarios / all_scenarios

END = time.time()

print(f"Created Foldy Spreadsheet in {round(END-START,2)} seconds.")

workbook.save(OUTPUT_FILE)
